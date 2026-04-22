"""
branch_analysis.py  —  Branch Supervision Data Analysis Tool
=============================================================
UCS SACCO  |  Risk & Compliance Department

Handles: Tills (any number), Treasury, Journals, Petty Cash
Works with any branch — all column detection is automatic.
Output matches Nkhatabay reference standard.

Run:   streamlit run branch_analysis.py
"""

import io, re, sys, os
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Branch Supervision Tool", layout="wide",
                   page_icon="", initial_sidebar_state="expanded")
st.markdown("""
<style>
[data-testid="stSidebar"]{background:#1A2942}
[data-testid="stSidebar"] *{color:#e8eef8!important}
[data-testid="metric-container"]{background:#f4f8fd;border:0.5px solid #d0d9e8;
  border-radius:8px;padding:12px 16px}
.block-container{padding-top:1rem}
.flag-critical{padding:10px 14px;border-radius:8px;margin:5px 0;font-size:13px;
  background:#fcebeb;border-left:4px solid #a32d2d;color:#791f1f}
.flag-high{padding:10px 14px;border-radius:8px;margin:5px 0;font-size:13px;
  background:#faeeda;border-left:4px solid #b45309;color:#633806}
.flag-info{padding:10px 14px;border-radius:8px;margin:5px 0;font-size:13px;
  background:#e6f1fb;border-left:4px solid #185fa5;color:#0c447c}
</style>""", unsafe_allow_html=True)

for k, v in {"branch":"","tills":{},"treasury":None,"journals":None,"petty":None}.items():
    if k not in st.session_state: st.session_state[k] = v

# ── helpers ────────────────────────────────────────────────────────────────────
def _num(s): return pd.to_numeric(s.astype(str).str.replace(",","").str.strip(),errors="coerce").fillna(0)
def _date(s): return pd.to_datetime(s,errors="coerce")
def _fmt(d):
    try: return pd.Timestamp(d).strftime("%d/%m/%Y")
    except: return str(d)
def flag(text,level="high"):
    cls={"critical":"flag-critical","high":"flag-high","info":"flag-info"}.get(level,"flag-info")
    st.markdown(f'<div class="{cls}">{text}</div>',unsafe_allow_html=True)
def show_df(df,height=420):
    d=df.copy()
    for c in d.columns:
        if pd.api.types.is_bool_dtype(d[c]): d[c]=d[c].map({True:"Yes",False:"No"})
    st.dataframe(d,width="stretch",height=height,hide_index=True)
def _member(det):
    for pat in [r"Withdrawn By:\s*([A-Z][A-Za-z\s]+?)(?:\s*[-\)]|$)",
                r"Deposited By:\s*([A-Z][A-Za-z\s]+?)(?:\s*[-\)]|$)",
                r"DIRECT RECEIPTS\(.*?-\s*Deposited By:\s*([A-Z][A-Za-z\s]+?)\)"]:
        m=re.search(pat,str(det),re.IGNORECASE)
        if m: return m.group(1).strip().title()
    return ""
def _cheque(det):
    m=re.search(r"CHEQUE\s*(?:NO\.?|NUMBER)?\s*(\d{3,6})",str(det).upper())
    return m.group(1) if m else ""
def _batch(det):
    m=re.search(r"batch\s*-\s*(\d+)",str(det),re.IGNORECASE)
    return m.group(1) if m else ""
def _voucher(det):
    m=re.search(r"Vno\.?\s*(\d{10,})",str(det),re.IGNORECASE)
    if not m: m=re.search(r"batch\s*-\s*(\d{10,})",str(det),re.IGNORECASE)
    return m.group(1) if m else ""

# ── cleaners ───────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def clean_till(fb,fn,label):
    eng="xlrd" if fn.lower().endswith(".xls") else "openpyxl"
    df=pd.read_excel(io.BytesIO(fb),engine=eng)
    df.columns=[str(c).strip().upper() for c in df.columns]
    if "DETAILS" in df.columns:
        df=df[~df["DETAILS"].astype(str).str.upper().str.contains("OPENING BALANCE",na=False)].copy()
    df["DATE"]=_date(df.get("DATE",pd.Series(dtype=str)))
    df["TIME_RAW"]=df.get("TIME",pd.Series(dtype=str)).astype(str)
    df["HOUR"]=df["TIME_RAW"].str.extract(r"^(\d{2}):",expand=False).astype(float)
    df["DEBIT"]=_num(df.get("DEBIT",pd.Series(0,index=df.index)))
    df["CREDIT"]=_num(df.get("CREDIT",pd.Series(0,index=df.index)))
    df["TILL"]=label; df["LINE_NO"]=range(1,len(df)+1)
    det=df.get("DETAILS",pd.Series("",index=df.index)).astype(str)
    def _cat(d):
        du=d.upper()
        if "REVERSAL" in du or du.startswith("REV-"): return "REVERSAL"
        if "CASH REQUEST" in du or "CCASHIER" in du: return "CASH_REQUEST"
        if "CASH WITHDRAWAL" in du or "CASH WITHDRAW" in du: return "WITHDRAWAL"
        if "CASH DEPOSIT" in du or "(CASH DEPOSIT)" in du: return "DEPOSIT"
        if any(x in du for x in ["DIRECT RECEIPT","LOAN FULL SETTLEMENT","LOAN PARTIAL",
                                   "FULL SETTLEMENT","PARTIAL PAYMENT"]): return "DIRECT_RECEIPT"
        if "TELLER" in du: return "TELLER_MOVEMENT"
        return "OTHER"
    df["CATEGORY"]=det.apply(_cat)
    df["MEMBER_NAME"]=det.apply(_member)
    df["VOUCHER_NO"]=det.apply(_voucher)
    df["DATE_FMT"]=df["DATE"].apply(_fmt)
    df["WEEKDAY"]=df["DATE"].dt.day_name()
    df["IS_WEEKEND"]=df["WEEKDAY"].isin(["Saturday","Sunday"])
    df["AFTER_HOURS"]=df["HOUR"].notna()&((df["HOUR"]<8)|(df["HOUR"]>=17))
    df["TIME_DISPLAY"]=df["TIME_RAW"].str[:8]
    df["BALANCE_DISPLAY"]=df.get("BALANCE",pd.Series("",index=df.index)).astype(str)
    return df

@st.cache_data(show_spinner=False)
def clean_treasury(fb,fn):
    eng="xlrd" if fn.lower().endswith(".xls") else "openpyxl"
    df=pd.read_excel(io.BytesIO(fb),engine=eng)
    df.columns=[str(c).strip().upper() for c in df.columns]
    df=df[~df.get("DETAILS",pd.Series("",index=df.index)).astype(str).str.upper()
           .str.contains("OPENING BALANCE",na=False)].copy()
    df["DATE"]=_date(df.get("DATE",pd.Series(dtype=str)))
    df["DATE_FMT"]=df["DATE"].apply(_fmt)
    df["TIME_DISPLAY"]=df.get("TIME",pd.Series("",index=df.index)).astype(str).str[:8]
    df["DEBIT"]=_num(df.get("DEBIT",pd.Series(0,index=df.index)))
    df["CREDIT"]=_num(df.get("CREDIT",pd.Series(0,index=df.index)))
    df["LINE_NO"]=range(1,len(df)+1)
    det=df.get("DETAILS",pd.Series("",index=df.index)).astype(str)
    def _cat(d):
        du=d.upper()
        if "REVERSAL" in du or du.strip().upper().startswith("REV-"): return "REVERSAL"
        if "CASH FROM BANK" in du or ("WITHDRAW FROM" in du and "NBM" in du) \
                or ("CHEQUE" in du and "TREASURY" in du): return "CASH_FROM_BANK"
        if "CASH TO BANK" in du or "CASH DEPOSIT FROM" in du \
                or ("DEPOSIT" in du and "NBM" in du): return "CASH_TO_BANK"
        if "CASH TO TELLER" in du: return "CASH_TO_TELLERS"
        if "CASH FROM TELLER" in du: return "CASH_FROM_TELLERS"
        if "PETTY" in du: return "PETTY_CASH_TOPUP"
        if "DIRECT RECEIPT" in du: return "DIRECT_RECEIPTS"
        if "OVERAGE" in du or "SHORTAGE" in du: return "CASH_DIFFERENCE"
        return "OTHER"
    df["CATEGORY"]=det.apply(_cat)
    df["CHEQUE_NO"]=det.apply(_cheque)
    df["BATCH_NO"]=det.apply(_batch)
    df["DEPOSITOR"]=det.apply(_member)
    def _act(d):
        d=str(d)
        d=re.sub(r"\s*batch\s*-\s*\d+","",d,flags=re.IGNORECASE)
        d=re.sub(r"^Journal\s*\(","",d,flags=re.IGNORECASE).rstrip(")").strip()
        return d.strip()
    df["ACTIVITY"]=det.apply(_act)
    df["BALANCE"]=df.get("BALANCE",pd.Series("",index=df.index)).astype(str)
    df["REFERENCE"]=df.get("REFERENCE",pd.Series("",index=df.index)).astype(str)
    return df

@st.cache_data(show_spinner=False)
def parse_journals(fb,fn):
    eng="xlrd" if fn.lower().endswith(".xls") else "openpyxl"
    raw=pd.read_excel(io.BytesIO(fb),engine=eng,header=None)
    records=[]; i=0; created_by=""; approved_by=""
    while i<len(raw):
        row=raw.iloc[i].tolist()
        if str(row[0]).strip()=="BatchNo":
            batch_no=str(row[1]).strip(); trdate=row[3]
            desc=str(row[5]).strip(); jtype=str(row[7]).strip() if pd.notna(row[7]) else ""
            created_by=""; approved_by=""; i+=2; legs=[]
            while i<len(raw):
                leg=raw.iloc[i].tolist(); c0=str(leg[0]).strip()
                if c0=="BatchNo": break
                if c0.startswith("Created By"):
                    created_by=str(leg[1]).strip()
                    approved_by=str(leg[5]).strip() if pd.notna(leg[5]) else ""
                    i+=1; break
                try:
                    debit=float(leg[5]) if pd.notna(leg[5]) else 0
                    credit=float(leg[6]) if pd.notna(leg[6]) else 0
                    legs.append({"BATCH_NO":batch_no,"DATE":trdate,"DESCRIPTION":desc,
                                 "JOURNAL_TYPE":jtype,
                                 "DR_ACCOUNT":str(leg[1]).strip(),"DR_NAME":str(leg[2]).strip(),
                                 "CR_ACCOUNT":str(leg[3]).strip(),"CR_NAME":str(leg[4]).strip(),
                                 "DEBIT":debit,"CREDIT":credit})
                except: pass
                i+=1
            for leg in legs:
                leg["CREATED_BY"]=created_by; leg["APPROVED_BY"]=approved_by
                records.append(leg)
        else: i+=1
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["DATE"]=_date(df["DATE"]); df["DATE_FMT"]=df["DATE"].apply(_fmt)
    df["WEEKDAY"]=df["DATE"].dt.day_name()
    df["IS_WEEKEND"]=df["WEEKDAY"].isin(["Saturday","Sunday"])
    df["AMOUNT"]=df["DEBIT"].where(df["DEBIT"]>0,df["CREDIT"])
    df["DESC_CLEAN"]=(df["DESCRIPTION"]
        .str.replace(r"^Journal\s*\(","",regex=True,flags=re.IGNORECASE)
        .str.rstrip(")").str.strip())
    def _cat(d):
        du=d.upper()
        if "CASH FROM BANK" in du or ("NBM" in du and "TREASURY" in du): return "CASH_FROM_BANK"
        if "CASH TO BANK" in du or "CASH DEPOSIT" in du: return "CASH_TO_BANK"
        if "TRANSPORT" in du: return "TRANSPORT"
        if "PETTY" in du: return "PETTY_CASH"
        if "MARKETING" in du or "BAM" in du or "BRING A MEMBER" in du: return "MARKETING_BAM"
        if "MEMBER EDUCATION" in du: return "MEMBER_EDUCATION"
        if "CIC" in du or "FUNERAL" in du or "INSURANCE" in du: return "INSURANCE"
        if "ELECTRICITY" in du or "WATER" in du or "BILL" in du: return "UTILITIES"
        if "SALARY" in du or "PAYROLL" in du: return "SALARY"
        if any(x in du for x in ["GROCERIES","CREMORA","SUGAR","MARGARINE","TOILET",
                                   "MOPPER","PRESTIGE"]): return "OFFICE_SUPPLIES"
        if "GENSET" in du or "FUEL" in du or "GENERATOR" in du: return "FUEL_GENERATOR"
        if "SOS" in du or "SAVINGS" in du: return "SAVINGS_TRANSFER"
        return "OTHER"
    df["CATEGORY"]=df["DESC_CLEAN"].apply(_cat)
    df["SAME_MAKER_CHECKER"]=df["CREATED_BY"].str.strip()==df["APPROVED_BY"].str.strip()
    df["NO_CHECKER"]=df["APPROVED_BY"].str.strip()==""
    SYS=["TREASURY","NBM","PETTY CASH","CCASHIER","TELLER","BRING A MEMBER",
         "MARKETING","MEMBER EDUCATION","CIC FUNERAL","CASH","BANK"]
    def _sys(n): return any(k in str(n).upper() for k in SYS)
    df["PERSON_TO_PERSON"]=(~df["DR_NAME"].apply(_sys)&~df["CR_NAME"].apply(_sys)&(df["DEBIT"]>0))
    df["SAME_PERSON_DR_CR"]=(df["DR_NAME"].str.strip().str.upper()==df["CR_NAME"].str.strip().str.upper())&(df["DEBIT"]>0)
    return df

@st.cache_data(show_spinner=False)
def clean_petty(fb,fn):
    eng="xlrd" if fn.lower().endswith(".xls") else "openpyxl"
    df=pd.read_excel(io.BytesIO(fb),engine=eng)
    df.columns=[str(c).strip().upper() for c in df.columns]
    df=df[~df.get("DETAILS",pd.Series("",index=df.index)).astype(str).str.upper()
           .str.contains("OPENING BALANCE",na=False)].copy()
    df["DATE"]=_date(df.get("DATE",pd.Series(dtype=str)))
    df["DATE_FMT"]=df["DATE"].apply(_fmt)
    df["TIME_DISPLAY"]=df.get("TIME",pd.Series("",index=df.index)).astype(str).str[:8]
    df["DEBIT"]=_num(df.get("DEBIT",pd.Series(0,index=df.index)))
    df["CREDIT"]=_num(df.get("CREDIT",pd.Series(0,index=df.index)))
    df["AMOUNT"]=df["DEBIT"].where(df["DEBIT"]>0,df["CREDIT"])
    df["LINE_NO"]=range(1,len(df)+1)
    det=df.get("DETAILS",pd.Series("",index=df.index)).astype(str)
    def _cat(d):
        du=d.upper()
        if "REVERSAL" in du or du.strip().upper().startswith("REV-"): return "REVERSAL"
        if "TOPUP" in du or "TOP UP" in du or "TOP-UP" in du: return "TOPUP"
        if "TRANSPORT" in du: return "TRANSPORT"
        if any(x in du for x in ["GROCERIES","CREMORA","SUGAR","MARGARINE","TOILET",
                                   "MOPPER","PRESTIGE"]): return "GROCERIES_SUPPLIES"
        if "ELECTRICITY" in du or "WATER" in du or "BILL" in du: return "UTILITIES"
        if "GENSET" in du or "FUEL" in du or "GENERATOR" in du: return "FUEL_GENERATOR"
        if "PRINTER" in du or "TONER" in du or "STATIONERY" in du: return "STATIONERY"
        return "OTHER"
    df["CATEGORY"]=det.apply(_cat)
    def _act(d):
        d=str(d)
        d=re.sub(r"\s*batch\s*-\s*\d+","",d,flags=re.IGNORECASE)
        d=re.sub(r"^Journal\s*\(","",d,flags=re.IGNORECASE).rstrip(")").strip()
        return d.strip()
    df["ACTIVITY"]=det.apply(_act)
    df["BATCH_NO"]=det.apply(_batch)
    df["BALANCE"]=df.get("BALANCE",pd.Series("",index=df.index)).astype(str)
    df["FLAG"]=""
    df.loc[df["CREDIT"]>=20000,"FLAG"]="Large payment ≥ MWK 20,000 — verify receipt and voucher"
    dup=df.duplicated(subset=["DATE","CREDIT","ACTIVITY"],keep=False)&(df["CREDIT"]>0)
    df.loc[dup&(df["FLAG"]==""),"FLAG"]="Duplicate: same activity + amount + date"
    return df

# ── excel builder ──────────────────────────────────────────────────────────────
NAVY="FF1A3A5C";GREEN="FF1E7E3E";AMBER="FFB45309";BLUE="FF185FA5"
RED="FFA32D2D";TEAL="FF0F6E56";PURPLE="FF534AB7";WHITE="FFFFFFFF";ALT="FFEEF4FB"

def _fmt_sheet(ws,hdr=NAVY):
    thin=Side(style="thin",color="FFD0D0D0"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    alt=PatternFill("solid",fgColor=ALT)
    for cell in ws[1]:
        cell.fill=PatternFill("solid",fgColor=hdr); cell.font=Font(bold=True,color=WHITE,size=9,name="Arial")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=bdr
    ws.row_dimensions[1].height=26; ws.freeze_panes="A2"
    for r,row in enumerate(ws.iter_rows(min_row=2),2):
        for cell in row:
            cell.border=bdr; cell.font=Font(size=9,name="Arial")
            if r%2==0:
                c=(cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "00000000")
                if c in ("FFFFFFFF","FF000000","00000000"): cell.fill=alt
    for col in ws.columns:
        ml=max((len(str(c.value or "")) for c in col),default=8)
        ws.column_dimensions[col[0].column_letter].width=min(max(ml+2,10),40)

def _safe(df):
    d=df.copy()
    for col in d.columns:
        if pd.api.types.is_bool_dtype(d[col]): d[col]=d[col].map({True:"Yes",False:"No"})
        elif d[col].dtype==object: d[col]=d[col].astype(str).replace("nan","")
    return d

def build_excel(branch,tills,treasury,journals,petty):
    buf=io.BytesIO(); sheets=[]
    def add(nm,df,col=NAVY):
        if df is not None and not df.empty: sheets.append((nm[:31],_safe(df),col))
    def _s(v):
        if isinstance(v,float): return f"MWK {v:,.2f}" if v>0 else "MWK 0.00"
        if isinstance(v,int): return f"{v:,}"
        return str(v)
    def _pick(df,cols): return df[[c for c in cols if c in df.columns]]

    all_till=pd.concat(tills.values(),ignore_index=True) if tills else pd.DataFrame()
    p0=all_till["DATE"].min() if not all_till.empty and "DATE" in all_till else None
    p1=all_till["DATE"].max() if not all_till.empty and "DATE" in all_till else None

    # SUMMARY
    rows=[("BRANCH SUPERVISION ANALYSIS REPORT",""),
          ("Branch",branch or "Not specified"),
          ("Period",f"{_fmt(p0)} to {_fmt(p1)}" if p0 else "N/A"),
          ("Generated",datetime.now().strftime("%d/%m/%Y %H:%M")),("","")]
    if not all_till.empty:
        wd=all_till[all_till["CATEGORY"]=="WITHDRAWAL"]
        dp=all_till[all_till["CATEGORY"]=="DEPOSIT"]
        dr=all_till[all_till["CATEGORY"]=="DIRECT_RECEIPT"]
        ah=all_till[all_till.get("AFTER_HOURS",pd.Series(False,index=all_till.index))==True] if "AFTER_HOURS" in all_till else pd.DataFrame()
        wk=all_till[all_till.get("IS_WEEKEND",pd.Series(False,index=all_till.index))==True] if "IS_WEEKEND" in all_till else pd.DataFrame()
        rows+=[("TILL SUMMARY",""),("Number of tills",_s(len(tills))),
               ("Total transactions (all tills)",_s(len(all_till))),
               ("Total cash withdrawals (MWK)",_s(wd["CREDIT"].sum())),
               ("Withdrawal count",_s(len(wd))),
               ("Total cash deposits (MWK)",_s(dp["DEBIT"].sum())),
               ("Total direct receipts — loan payments (MWK)",_s(dr["DEBIT"].sum())),
               ("After-hours transactions — FINDING",_s(len(ah))),
               ("Weekend transactions — FINDING",_s(len(wk))),("","")]
    if treasury is not None and not treasury.empty:
        cfb=treasury[treasury["CATEGORY"]=="CASH_FROM_BANK"]
        ctb=treasury[treasury["CATEGORY"]=="CASH_TO_BANK"]
        ctt=treasury[treasury["CATEGORY"]=="CASH_TO_TELLERS"]
        rows+=[("TREASURY SUMMARY",""),
               ("Cash received from bank (MWK)",_s(cfb["DEBIT"].sum())),
               ("Number of cheque withdrawals",_s(len(cfb))),
               ("Cash deposited to bank (MWK)",_s(ctb["CREDIT"].sum())),
               ("Cash issued to tellers (MWK)",_s(ctt["CREDIT"].sum())),("","")]
    if journals is not None and not journals.empty:
        sm=int(journals["SAME_MAKER_CHECKER"].sum()); nc=int(journals["NO_CHECKER"].sum())
        p2=int(journals["PERSON_TO_PERSON"].sum()); sd=int(journals["SAME_PERSON_DR_CR"].sum())
        rows+=[("JOURNAL SUMMARY",""),
               ("Total journal batches",_s(journals["BATCH_NO"].nunique())),
               ("Total journal legs",_s(len(journals))),
               ("Total debited (MWK)",_s(journals["DEBIT"].sum())),
               ("Same maker and checker — CRITICAL FINDING",_s(sm)),
               ("No checker recorded — CRITICAL FINDING",_s(nc)),
               ("Person-to-person transfers — Investigate",_s(p2)),
               ("Same person DR and CR — Investigate",_s(sd)),("","")]
    if petty is not None and not petty.empty:
        rows+=[("PETTY CASH SUMMARY",""),
               ("Total petty transactions",_s(len(petty))),
               ("Total expenditure (MWK)",_s(petty["CREDIT"].sum())),
               ("Anomaly flags",_s(int((petty["FLAG"]!="").sum())))]
    add("SUMMARY",pd.DataFrame(rows,columns=["Item","Value"]),NAVY)

    # TILLS
    WALL=["LINE_NO","DATE_FMT","TIME_DISPLAY","CATEGORY","MEMBER_NAME","VOUCHER_NO",
          "DEBIT","CREDIT","BALANCE_DISPLAY","REFERENCE","DETAILS"]
    for tname,tdf in tills.items():
        lbl=tname.upper().replace(" ","_").replace(".","")
        wd=tdf[tdf["CATEGORY"]=="WITHDRAWAL"]; dp=tdf[tdf["CATEGORY"]=="DEPOSIT"]
        dr=tdf[tdf["CATEGORY"]=="DIRECT_RECEIPT"]; cr=tdf[tdf["CATEGORY"]=="CASH_REQUEST"]
        ah=tdf[tdf.get("AFTER_HOURS",pd.Series(False,index=tdf.index))==True] if "AFTER_HOURS" in tdf else pd.DataFrame()
        wk=tdf[tdf.get("IS_WEEKEND",pd.Series(False,index=tdf.index))==True] if "IS_WEEKEND" in tdf else pd.DataFrame()
        ts=[("Till",tname),("Total transactions",_s(len(tdf))),
            ("Cash withdrawals (MWK)",_s(wd["CREDIT"].sum())),("Withdrawal count",_s(len(wd))),
            ("Cash deposits received (MWK)",_s(dp["DEBIT"].sum())),("Deposit count",_s(len(dp))),
            ("Direct receipts — loan payments (MWK)",_s(dr["DEBIT"].sum())),("Direct receipt count",_s(len(dr))),
            ("Cash requests from CCASHIER (MWK)",_s(cr["DEBIT"].sum())),
            ("After-hours transactions",_s(len(ah))),("Weekend transactions",_s(len(wk)))]
        add(f"{lbl}_SUMMARY",pd.DataFrame(ts,columns=["Item","Value"]),GREEN)

        WD_COLS=["LINE_NO","DATE_FMT","TIME_DISPLAY","WEEKDAY","MEMBER_NAME","VOUCHER_NO",
                 "CREDIT","BALANCE_DISPLAY","AFTER_HOURS","REFERENCE"]
        if not wd.empty:
            w=_pick(wd.sort_values("DATE"),WD_COLS)
            w=w.rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","CREDIT":"AMOUNT_WITHDRAWN","BALANCE_DISPLAY":"BALANCE"})
            add(f"{lbl}_WITHDRAWALS",w,GREEN)

        DEP_COLS=["LINE_NO","DATE_FMT","TIME_DISPLAY","MEMBER_NAME","VOUCHER_NO","DEBIT","BALANCE_DISPLAY","REFERENCE"]
        if not dp.empty:
            d=_pick(dp.sort_values("DATE"),DEP_COLS)
            d=d.rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","DEBIT":"AMOUNT_DEPOSITED","BALANCE_DISPLAY":"BALANCE"})
            add(f"{lbl}_DEPOSITS",d,TEAL)

        DR_COLS=["LINE_NO","DATE_FMT","TIME_DISPLAY","MEMBER_NAME","VOUCHER_NO","DEBIT","BALANCE_DISPLAY","REFERENCE"]
        if not dr.empty:
            r=_pick(dr.sort_values("DATE"),DR_COLS)
            r=r.rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","DEBIT":"AMOUNT_RECEIVED","BALANCE_DISPLAY":"BALANCE"})
            add(f"{lbl}_DIRECT_RECEIPTS",r,BLUE)

        daily=(tdf.groupby("DATE_FMT",sort=False)
               .agg(DATE=("DATE","first"),
                    WITHDRAWALS=("CATEGORY",lambda x:(x=="WITHDRAWAL").sum()),
                    TOTAL_WITHDRAWN=("CREDIT",lambda x:x[tdf.loc[x.index,"CATEGORY"]=="WITHDRAWAL"].sum()),
                    DEPOSITS=("CATEGORY",lambda x:(x=="DEPOSIT").sum()),
                    TOTAL_DEPOSITED=("DEBIT",lambda x:x[tdf.loc[x.index,"CATEGORY"]=="DEPOSIT"].sum()),
                    DIRECT_RECEIPTS=("CATEGORY",lambda x:(x=="DIRECT_RECEIPT").sum()),
                    TOTAL_DR=("DEBIT",lambda x:x[tdf.loc[x.index,"CATEGORY"]=="DIRECT_RECEIPT"].sum()))
               .reset_index(drop=True).sort_values("DATE"))
        daily["DATE"]=daily["DATE"].apply(_fmt)
        add(f"{lbl}_DAILY_SUMMARY",daily,NAVY)

        if "MEMBER_NAME" in wd.columns and not wd.empty:
            ms=(wd[wd["MEMBER_NAME"]!=""].groupby("MEMBER_NAME")
                .agg(WITHDRAWALS=("CREDIT","count"),TOTAL_WITHDRAWN=("CREDIT","sum"),LARGEST_SINGLE=("CREDIT","max"))
                .reset_index().sort_values("TOTAL_WITHDRAWN",ascending=False))
            add(f"{lbl}_MEMBER_SUMMARY",ms,PURPLE)

        all_out=_pick(tdf.sort_values("DATE"),WALL)
        all_out=all_out.rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","BALANCE_DISPLAY":"BALANCE"})
        add(f"{lbl}_ALL_TRANSACTIONS",all_out,NAVY)

        anoms=[]
        if not ah.empty:
            a=_pick(ah,["LINE_NO","DATE_FMT","TIME_DISPLAY","CATEGORY","MEMBER_NAME","CREDIT","DEBIT"]).copy()
            a["FLAG"]="After-hours transaction"; a=a.rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}); anoms.append(a)
        if not wk.empty:
            a=_pick(wk,["LINE_NO","DATE_FMT","WEEKDAY","CATEGORY","MEMBER_NAME","CREDIT","DEBIT"]).copy()
            a["FLAG"]="Weekend transaction"; a=a.rename(columns={"DATE_FMT":"DATE"}); anoms.append(a)
        lrg=wd[wd["CREDIT"]>=1_000_000]
        if not lrg.empty:
            a=_pick(lrg,["LINE_NO","DATE_FMT","MEMBER_NAME","CREDIT"]).copy()
            a["FLAG"]="Single withdrawal ≥ MWK 1,000,000 — verify authorisation"
            a=a.rename(columns={"DATE_FMT":"DATE","CREDIT":"AMOUNT"}); anoms.append(a)
        dup_m=wd.duplicated(subset=["DATE_FMT","MEMBER_NAME","CREDIT"],keep=False)&(wd["MEMBER_NAME"]!="")&(wd["CREDIT"]>0)
        dup_wd=wd[dup_m]
        if not dup_wd.empty:
            a=_pick(dup_wd,["LINE_NO","DATE_FMT","MEMBER_NAME","CREDIT"]).copy()
            a["FLAG"]="Duplicate: same member + amount + date"; a=a.rename(columns={"DATE_FMT":"DATE","CREDIT":"AMOUNT"}); anoms.append(a)
        if anoms: add(f"{lbl}_ANOMALIES",pd.concat(anoms,ignore_index=True),RED)

    # TREASURY
    if treasury is not None and not treasury.empty:
        TR=["LINE_NO","DATE_FMT","TIME_DISPLAY","CHEQUE_NO","BATCH_NO","ACTIVITY","DEBIT","CREDIT","BALANCE","REFERENCE"]
        tcat=(treasury.groupby("CATEGORY").agg(ROWS=("DEBIT","count"),TOTAL_DEBIT=("DEBIT","sum"),TOTAL_CREDIT=("CREDIT","sum")).reset_index())
        tcat["NET"]=tcat["TOTAL_DEBIT"]-tcat["TOTAL_CREDIT"]
        add("TREASURY_SUMMARY",tcat,NAVY)
        cfb=treasury[treasury["CATEGORY"]=="CASH_FROM_BANK"]
        if not cfb.empty:
            chq=cfb[["DATE_FMT","CHEQUE_NO","BATCH_NO","DEBIT","ACTIVITY","REFERENCE"]].copy()
            chq.columns=["DATE","CHEQUE_NO","BATCH_NO","AMOUNT_DRAWN","DESCRIPTION","REFERENCE"]
            chq=chq.sort_values("CHEQUE_NO"); chq["VERIFIED"]=""
            add("CHEQUE_REGISTER",chq,GREEN)
            nums=pd.to_numeric(chq["CHEQUE_NO"],errors="coerce").dropna().astype(int).sort_values()
            if len(nums)>1:
                missing=sorted(set(range(nums.min(),nums.max()+1))-set(nums))
                if missing:
                    add("CHEQUE_GAPS",pd.DataFrame({"CHEQUE_NO":missing,"STATUS":"NOT IN SYSTEM DATA",
                        "ACTION_REQUIRED":"Verify in physical cheque book — used, cancelled, or returned?"}),RED)
        for cat,nm,col in [("CASH_FROM_BANK","CASH_FROM_BANK",GREEN),("CASH_TO_BANK","CASH_TO_BANK",BLUE),
                            ("CASH_TO_TELLERS","CASH_TO_TELLERS",TEAL),("CASH_FROM_TELLERS","CASH_FROM_TELLERS",TEAL),
                            ("PETTY_CASH_TOPUP","PETTY_TOPUP",PURPLE),("REVERSAL","TREASURY_REVERSALS",RED),
                            ("CASH_DIFFERENCE","CASH_DIFFERENCES",AMBER),("OTHER","TREASURY_OTHER",NAVY)]:
            sub=treasury[treasury["CATEGORY"]==cat]
            if not sub.empty:
                add(nm,_pick(sub,TR).rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}),col)

    # JOURNALS
    if journals is not None and not journals.empty:
        JC=["BATCH_NO","DATE_FMT","WEEKDAY","DESC_CLEAN","CATEGORY","CREATED_BY","APPROVED_BY","DR_NAME","CR_NAME","DEBIT","CREDIT"]
        def jout(df): return _pick(df,JC).rename(columns={"DATE_FMT":"DATE"})
        add("JOURNALS_ALL",jout(journals.sort_values("DATE")),BLUE)
        for cat,nm,col in [("CASH_FROM_BANK","J_CASH_FROM_BANK",GREEN),("CASH_TO_BANK","J_CASH_TO_BANK",TEAL),
                            ("MARKETING_BAM","J_MARKETING_BAM",AMBER),("TRANSPORT","J_TRANSPORT",NAVY),
                            ("PETTY_CASH","J_PETTY_CASH",PURPLE),("SAVINGS_TRANSFER","J_SAVINGS_TRANSFER",BLUE),
                            ("MEMBER_EDUCATION","J_MEMBER_EDUCATION",TEAL),("INSURANCE","J_INSURANCE",NAVY),
                            ("UTILITIES","J_UTILITIES",NAVY),("OFFICE_SUPPLIES","J_OFFICE_SUPPLIES",NAVY),
                            ("FUEL_GENERATOR","J_FUEL_GENERATOR",AMBER)]:
            sub=journals[journals["CATEGORY"]==cat]
            if not sub.empty: add(nm,jout(sub.sort_values("DATE")),col)

        inv=[]
        for df_sub,msg in [(journals[journals["SAME_MAKER_CHECKER"]==True],"SAME MAKER AND CHECKER — control failure"),
                           (journals[journals["NO_CHECKER"]==True],"NO APPROVER RECORDED"),
                           (journals[journals["PERSON_TO_PERSON"]==True],"PERSON-TO-PERSON TRANSFER — verify legitimacy"),
                           (journals[journals["SAME_PERSON_DR_CR"]==True],"SAME PERSON DEBIT AND CREDIT — investigate"),
                           (journals[journals["IS_WEEKEND"]==True].drop_duplicates("BATCH_NO"),"WEEKEND JOURNAL — verify physical authorisation")]:
            for _,r in df_sub.iterrows():
                d={c:r[c] for c in JC if c in r.index}; d["FINDING"]=msg; inv.append(d)
        if inv:
            idf=pd.DataFrame(inv)
            if "DATE_FMT" in idf.columns: idf=idf.rename(columns={"DATE_FMT":"DATE"})
            add("JOURNALS_INVESTIGATE",idf,RED)

        add("J_CREATORS",journals.groupby("CREATED_BY").agg(BATCHES=("BATCH_NO","nunique"),TOTAL_DEBIT=("DEBIT","sum")).reset_index().sort_values("TOTAL_DEBIT",ascending=False),NAVY)
        add("J_APPROVERS",journals[journals["APPROVED_BY"].str.strip()!=""].groupby("APPROVED_BY").agg(APPROVED=("BATCH_NO","nunique")).reset_index().sort_values("APPROVED",ascending=False),NAVY)

    # PETTY
    if petty is not None and not petty.empty:
        pc=["LINE_NO","DATE_FMT","TIME_DISPLAY","BATCH_NO","CATEGORY","ACTIVITY","DEBIT","CREDIT","BALANCE","FLAG"]
        reg=_pick(petty,pc).rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}); reg["VERIFIED"]=""
        add("PETTY_REGISTER",reg,TEAL)
        cs=(petty[petty["CREDIT"]>0].groupby("CATEGORY").agg(TRANSACTIONS=("CREDIT","count"),TOTAL_SPENT=("CREDIT","sum")).reset_index().sort_values("TOTAL_SPENT",ascending=False))
        tot=cs["TOTAL_SPENT"].sum(); cs["PCT_OF_TOTAL"]=(cs["TOTAL_SPENT"]/tot*100).round(1)
        add("PETTY_CATEGORY_SUMMARY",cs,NAVY)
        dp2=(petty.groupby("DATE_FMT",sort=False).agg(DATE=("DATE","first"),TRANSACTIONS=("CREDIT","count"),TOTAL_SPENT=("CREDIT","sum")).reset_index(drop=True).sort_values("DATE"))
        dp2["DATE"]=dp2["DATE"].apply(_fmt); dp2["VERIFIED"]=""
        add("PETTY_DAILY_SUMMARY",dp2,NAVY)
        ap=petty[petty["FLAG"]!=""][[c for c in ["LINE_NO","DATE_FMT","ACTIVITY","CREDIT","FLAG"] if c in petty.columns]].rename(columns={"DATE_FMT":"DATE","CREDIT":"AMOUNT"})
        if not ap.empty: add("PETTY_ANOMALIES",ap,RED)

    # write
    with pd.ExcelWriter(buf,engine="openpyxl") as writer:
        for nm,df,_ in sheets: df.to_excel(writer,sheet_name=nm,index=False)
    buf.seek(0); wb=load_workbook(buf)
    for nm,_,col in sheets:
        if nm in wb.sheetnames: _fmt_sheet(wb[nm],col)
    out=io.BytesIO(); wb.save(out); out.seek(0); return out.read()

# ── sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### Branch Supervision")
    st.markdown("**Data Analysis Tool**")
    st.divider()
    page=st.radio("Navigate",["Upload & Clean","Till Analysis","Treasury",
                               "Journals & Fraud","Petty Cash","Export Report"],
                  label_visibility="collapsed")
    st.divider()
    st.markdown("**Session**")
    st.markdown(f"Branch: `{st.session_state.branch or 'not set'}`")
    st.markdown(f"Tills: {len(st.session_state.tills)} loaded")
    st.markdown(f"Treasury: {'yes' if st.session_state.treasury is not None else 'no'}")
    jdf=st.session_state.journals
    st.markdown(f"Journals: {'yes' if jdf is not None and not jdf.empty else 'no'}")
    st.markdown(f"Petty: {'yes' if st.session_state.petty is not None else 'no'}")
    st.divider()
    if st.button("Reset session",use_container_width=True):
        for k in list(st.session_state.keys()): del st.session_state[k]
        st.rerun()
    st.caption("Risk & Compliance Dept")

# ── pages ──────────────────────────────────────────────────────────────────────
if page=="Upload & Clean":
    st.title("Upload & Clean Branch Files")
    st.caption("Column detection is automatic — works with any branch.")
    cb,_=st.columns([1,3])
    with cb:
        b=st.text_input("Branch name",value=st.session_state.branch,placeholder="e.g. Mzimba")
        if b: st.session_state.branch=b
    st.divider()
    c1,c2,c3,c4=st.columns(4)
    with c1:
        st.subheader("Till files")
        tfs=st.file_uploader("Till files",type=["xls","xlsx"],accept_multiple_files=True,label_visibility="collapsed")
        if tfs:
            for f in tfs:
                ln=f.name.lower(); m=re.search(r"till[_\s]*(\d+)",ln)
                lbl=f"Till {m.group(1)}" if m else f.name.split(".")[0].title()
                with st.spinner(f"Cleaning {f.name}..."):
                    st.session_state.tills[lbl]=clean_till(f.read(),f.name,lbl)
            st.success(f"{len(st.session_state.tills)} till(s) loaded.")
    with c2:
        st.subheader("Treasury file")
        tf=st.file_uploader("Treasury",type=["xls","xlsx"],label_visibility="collapsed")
        if tf:
            with st.spinner("Cleaning..."):
                st.session_state.treasury=clean_treasury(tf.read(),tf.name)
            st.success(f"Treasury: {len(st.session_state.treasury):,} rows")
    with c3:
        st.subheader("Journal file")
        jf=st.file_uploader("Journals",type=["xls","xlsx"],label_visibility="collapsed")
        if jf:
            with st.spinner("Parsing..."):
                j=parse_journals(jf.read(),jf.name); st.session_state.journals=j
            if j is not None and not j.empty:
                st.success(f"Journals: {j['BATCH_NO'].nunique():,} batches")
            else: st.error("Could not parse — check format.")
    with c4:
        st.subheader("Petty cash file")
        pf=st.file_uploader("Petty",type=["xls","xlsx"],label_visibility="collapsed")
        if pf:
            with st.spinner("Cleaning..."):
                st.session_state.petty=clean_petty(pf.read(),pf.name)
            st.success(f"Petty: {len(st.session_state.petty):,} rows")
    if st.session_state.tills or st.session_state.treasury is not None:
        st.divider(); st.info("Files loaded. Use the sidebar to navigate.")

elif page=="Till Analysis":
    st.title("Till Transaction Analysis")
    tills=st.session_state.tills
    if not tills: st.warning("No till files loaded."); st.stop()
    all_till=pd.concat(tills.values(),ignore_index=True)
    wd=all_till[all_till["CATEGORY"]=="WITHDRAWAL"]; dp=all_till[all_till["CATEGORY"]=="DEPOSIT"]
    dr=all_till[all_till["CATEGORY"]=="DIRECT_RECEIPT"]
    ah=int(all_till.get("AFTER_HOURS",pd.Series(False)).sum()) if "AFTER_HOURS" in all_till else 0
    wk=int(all_till.get("IS_WEEKEND",pd.Series(False)).sum()) if "IS_WEEKEND" in all_till else 0
    c1,c2,c3,c4,c5=st.columns(5)
    with c1: st.metric("Transactions",f"{len(all_till):,}")
    with c2: st.metric("Withdrawals (MWK)",f"{wd['CREDIT'].sum():,.0f}")
    with c3: st.metric("Deposits (MWK)",f"{dp['DEBIT'].sum():,.0f}")
    with c4: st.metric("Direct receipts (MWK)",f"{dr['DEBIT'].sum():,.0f}")
    with c5: st.metric("After-hours + Weekend",f"{ah+wk}",delta="Investigate" if ah+wk>0 else None,delta_color="inverse")
    st.divider()
    sel=st.selectbox("View",["All tills combined"]+list(tills.keys()))
    vdf=tills[sel] if sel!="All tills combined" else all_till
    t1,t2,t3,t4,t5=st.tabs(["Withdrawals","Deposits","Direct Receipts","Member Summary","Anomalies"])
    with t1:
        w=vdf[vdf["CATEGORY"]=="WITHDRAWAL"]
        st.caption(f"{len(w):,} withdrawals — MWK {w['CREDIT'].sum():,.0f}")
        if not w.empty:
            show_df(w[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","MEMBER_NAME","VOUCHER_NO","CREDIT","AFTER_HOURS","IS_WEEKEND","BALANCE_DISPLAY"] if c in w.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","CREDIT":"AMOUNT_WITHDRAWN","BALANCE_DISPLAY":"BALANCE"}))
        if "DATE" in w.columns and not w.empty:
            d2=w.groupby(w["DATE"].dt.date)["CREDIT"].sum().reset_index(); d2.columns=["Date","Amount"]
            fig=px.bar(d2,x="Date",y="Amount",title="Daily total withdrawals (MWK)",height=260)
            fig.update_layout(showlegend=False); st.plotly_chart(fig,use_container_width=True)
    with t2:
        d=vdf[vdf["CATEGORY"]=="DEPOSIT"]
        st.caption(f"{len(d):,} deposits — MWK {d['DEBIT'].sum():,.0f}")
        if not d.empty:
            show_df(d[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","MEMBER_NAME","VOUCHER_NO","DEBIT","BALANCE_DISPLAY"] if c in d.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","DEBIT":"AMOUNT_DEPOSITED","BALANCE_DISPLAY":"BALANCE"}))
    with t3:
        r=vdf[vdf["CATEGORY"]=="DIRECT_RECEIPT"]
        st.caption(f"{len(r):,} direct receipts — MWK {r['DEBIT'].sum():,.0f}")
        if not r.empty:
            show_df(r[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","MEMBER_NAME","VOUCHER_NO","DEBIT","BALANCE_DISPLAY"] if c in r.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME","DEBIT":"AMOUNT_RECEIVED","BALANCE_DISPLAY":"BALANCE"}))
    with t4:
        w2=vdf[vdf["CATEGORY"]=="WITHDRAWAL"]
        if "MEMBER_NAME" in w2.columns and not w2.empty:
            ms=(w2[w2["MEMBER_NAME"]!=""].groupby("MEMBER_NAME").agg(WITHDRAWALS=("CREDIT","count"),TOTAL_WITHDRAWN=("CREDIT","sum"),LARGEST_SINGLE=("CREDIT","max")).reset_index().sort_values("TOTAL_WITHDRAWN",ascending=False))
            show_df(ms.head(50),height=500)
            fig=px.bar(ms.head(15),x="MEMBER_NAME",y="TOTAL_WITHDRAWN",title="Top 15 members (MWK)",height=280)
            fig.update_layout(xaxis_tickangle=-40); st.plotly_chart(fig,use_container_width=True)
    with t5:
        ah_df=vdf[vdf.get("AFTER_HOURS",pd.Series(False,index=vdf.index))==True] if "AFTER_HOURS" in vdf else pd.DataFrame()
        wk_df=vdf[vdf.get("IS_WEEKEND",pd.Series(False,index=vdf.index))==True] if "IS_WEEKEND" in vdf else pd.DataFrame()
        if not ah_df.empty:
            flag(f"{len(ah_df)} after-hours transaction(s).","high")
            show_df(ah_df[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","CATEGORY","MEMBER_NAME","CREDIT","DEBIT"] if c in ah_df.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}),height=250)
        if not wk_df.empty:
            flag(f"{len(wk_df)} weekend transaction(s).","high")
            show_df(wk_df[[c for c in ["LINE_NO","DATE_FMT","WEEKDAY","CATEGORY","MEMBER_NAME","CREDIT","DEBIT"] if c in wk_df.columns]].rename(columns={"DATE_FMT":"DATE"}),height=250)
        if ah_df.empty and wk_df.empty: st.success("No anomalies.")

elif page=="Treasury":
    st.title("Treasury Analysis")
    tr=st.session_state.treasury
    if tr is None: st.warning("No treasury file loaded."); st.stop()
    cfb=tr[tr["CATEGORY"]=="CASH_FROM_BANK"]; ctb=tr[tr["CATEGORY"]=="CASH_TO_BANK"]
    ctt=tr[tr["CATEGORY"]=="CASH_TO_TELLERS"]; cft=tr[tr["CATEGORY"]=="CASH_FROM_TELLERS"]
    c1,c2,c3,c4=st.columns(4)
    with c1: st.metric("Cash from bank",f"MWK {cfb['DEBIT'].sum():,.0f}")
    with c2: st.metric("Cash to bank",f"MWK {ctb['CREDIT'].sum():,.0f}")
    with c3: st.metric("Issued to tellers",f"MWK {ctt['CREDIT'].sum():,.0f}")
    with c4: st.metric("Received from tellers",f"MWK {cft['DEBIT'].sum():,.0f}")
    st.divider()
    t1,t2,t3=st.tabs(["Cheque Register","By Category","All Transactions"])
    with t1:
        if not cfb.empty:
            chq=cfb[["DATE_FMT","CHEQUE_NO","BATCH_NO","DEBIT","ACTIVITY"]].rename(columns={"DATE_FMT":"DATE","DEBIT":"AMOUNT_DRAWN","ACTIVITY":"DESCRIPTION"}); chq["VERIFIED"]=""
            show_df(chq.sort_values("CHEQUE_NO"),height=500)
            nums=pd.to_numeric(chq["CHEQUE_NO"],errors="coerce").dropna().astype(int)
            if len(nums)>1:
                miss=sorted(set(range(nums.min(),nums.max()+1))-set(nums))
                if miss: flag(f"Cheque gaps: {miss[:20]}{'...' if len(miss)>20 else ''}. Verify against physical cheque book.","high")
        else: st.info("No CASH_FROM_BANK entries found.")
    with t2:
        cat=st.selectbox("Category",sorted(tr["CATEGORY"].unique()))
        sub=tr[tr["CATEGORY"]==cat]
        show_df(sub[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","CHEQUE_NO","ACTIVITY","DEBIT","CREDIT","BALANCE","REFERENCE"] if c in sub.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}))
    with t3:
        show_df(tr[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","CATEGORY","CHEQUE_NO","ACTIVITY","DEBIT","CREDIT","BALANCE"] if c in tr.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}))

elif page=="Journals & Fraud":
    st.title("Journal Analysis & Fraud Detection")
    jdf=st.session_state.journals
    if jdf is None or jdf.empty: st.warning("No journal file loaded."); st.stop()
    sm=int(jdf["SAME_MAKER_CHECKER"].sum()); nc=int(jdf["NO_CHECKER"].sum())
    p2=int(jdf["PERSON_TO_PERSON"].sum()); sd=int(jdf["SAME_PERSON_DR_CR"].sum()); wk=int(jdf["IS_WEEKEND"].sum())
    c1,c2,c3,c4,c5=st.columns(5)
    with c1: st.metric("Batches",f"{jdf['BATCH_NO'].nunique():,}")
    with c2: st.metric("Same maker/checker",f"{sm}",delta="CRITICAL" if sm>0 else None,delta_color="inverse")
    with c3: st.metric("No checker",f"{nc}",delta="CRITICAL" if nc>0 else None,delta_color="inverse")
    with c4: st.metric("Person-to-person",f"{p2}",delta="Investigate" if p2>0 else None,delta_color="inverse")
    with c5: st.metric("Weekend journals",f"{wk}",delta="Review" if wk>0 else None,delta_color="inverse")
    st.divider()
    JD=["BATCH_NO","DATE_FMT","WEEKDAY","DESC_CLEAN","CATEGORY","CREATED_BY","APPROVED_BY","DR_NAME","CR_NAME","DEBIT","CREDIT"]
    def jshow(df): show_df(_pick(df,[c for c in JD if c in df.columns]).rename(columns={"DATE_FMT":"DATE"}))
    def _pick(df,cols): return df[[c for c in cols if c in df.columns]]
    t1,t2,t3,t4=st.tabs(["All journals","Fraud & Investigate","By category","By person"])
    with t1: jshow(jdf.sort_values("DATE"))
    with t2:
        if sm>0:
            flag(f"MAKER-CHECKER VIOLATION: {sm} leg(s). MWK {jdf[jdf['SAME_MAKER_CHECKER']==True]['DEBIT'].sum():,.0f}","critical")
            jshow(jdf[jdf["SAME_MAKER_CHECKER"]==True].drop_duplicates("BATCH_NO")); st.divider()
        if nc>0:
            flag(f"NO APPROVER: {nc} leg(s) have no checker recorded.","critical")
            jshow(jdf[jdf["NO_CHECKER"]==True].drop_duplicates("BATCH_NO")); st.divider()
        if p2>0:
            flag(f"PERSON-TO-PERSON: {p2} leg(s) transfer between individuals. Verify each.","high")
            jshow(jdf[jdf["PERSON_TO_PERSON"]==True]); st.divider()
        if sd>0:
            flag(f"SAME PERSON DR & CR: {sd} leg(s). Investigate.","high")
            jshow(jdf[jdf["SAME_PERSON_DR_CR"]==True]); st.divider()
        if wk>0:
            flag(f"WEEKEND JOURNALS: {wk} entries.","high")
            jshow(jdf[jdf["IS_WEEKEND"]==True].drop_duplicates("BATCH_NO"))
        if sm==0 and nc==0 and p2==0 and sd==0 and wk==0: st.success("No fraud flags.")
    with t3:
        cat=(jdf[jdf["DEBIT"]>0].groupby("CATEGORY").agg(BATCHES=("BATCH_NO","nunique"),TOTAL_MWK=("DEBIT","sum")).reset_index().sort_values("TOTAL_MWK",ascending=False))
        c1,c2=st.columns([1,2])
        with c1: show_df(cat,height=350)
        with c2:
            fig=px.pie(cat,values="TOTAL_MWK",names="CATEGORY",hole=0.45,title="Journal amounts by category",height=300)
            st.plotly_chart(fig,use_container_width=True)
        sc=st.selectbox("Drill into category",sorted(jdf["CATEGORY"].unique()))
        jshow(jdf[jdf["CATEGORY"]==sc].sort_values("DATE"))
    with t4:
        cc,ca=st.columns(2)
        with cc:
            st.caption("Created by")
            show_df(jdf.groupby("CREATED_BY").agg(BATCHES=("BATCH_NO","nunique"),TOTAL_DEBIT=("DEBIT","sum")).reset_index().sort_values("TOTAL_DEBIT",ascending=False),height=300)
        with ca:
            st.caption("Approved by")
            ap=jdf[jdf["APPROVED_BY"].str.strip()!=""]
            show_df(ap.groupby("APPROVED_BY").agg(APPROVED=("BATCH_NO","nunique")).reset_index().sort_values("APPROVED",ascending=False),height=300)

elif page=="Petty Cash":
    st.title("Petty Cash Analysis")
    pdf=st.session_state.petty
    if pdf is None or pdf.empty: st.warning("No petty cash file loaded."); st.stop()
    c1,c2,c3=st.columns(3)
    with c1: st.metric("Transactions",f"{len(pdf):,}")
    with c2: st.metric("Total expenditure (MWK)",f"{pdf['CREDIT'].sum():,.0f}")
    with c3: st.metric("Anomaly flags",f"{int((pdf['FLAG']!='').sum())}",delta="Review" if (pdf['FLAG']!='').sum()>0 else None,delta_color="inverse")
    st.divider()
    pt1,pt2,pt3=st.tabs(["Transaction register","Category summary","Anomalies"])
    with pt1:
        show_df(pdf[[c for c in ["LINE_NO","DATE_FMT","TIME_DISPLAY","CATEGORY","ACTIVITY","DEBIT","CREDIT","BALANCE","FLAG"] if c in pdf.columns]].rename(columns={"DATE_FMT":"DATE","TIME_DISPLAY":"TIME"}))
    with pt2:
        cs=(pdf[pdf["CREDIT"]>0].groupby("CATEGORY").agg(TRANSACTIONS=("CREDIT","count"),TOTAL_SPENT=("CREDIT","sum")).reset_index().sort_values("TOTAL_SPENT",ascending=False))
        tot=cs["TOTAL_SPENT"].sum(); cs["PCT"]=(cs["TOTAL_SPENT"]/tot*100).round(1)
        c1,c2=st.columns([1,2])
        with c1: show_df(cs,height=300)
        with c2:
            fig=px.pie(cs,values="TOTAL_SPENT",names="CATEGORY",hole=0.45,title="Petty cash by category",height=280)
            st.plotly_chart(fig,use_container_width=True)
    with pt3:
        ap=pdf[pdf["FLAG"]!=""]
        if ap.empty: st.success("No anomalies.")
        else: show_df(ap[[c for c in ["LINE_NO","DATE_FMT","ACTIVITY","CREDIT","FLAG"] if c in ap.columns]].rename(columns={"DATE_FMT":"DATE","CREDIT":"AMOUNT"}))

elif page=="Export Report":
    st.title("Export Full Analysis Report")
    tills=st.session_state.tills; treasury=st.session_state.treasury
    journals=st.session_state.journals; petty=st.session_state.petty
    branch=st.session_state.branch or "BRANCH"
    if not tills and treasury is None and journals is None and petty is None:
        st.warning("No data loaded."); st.stop()
    if st.button("Generate workbook",type="primary"):
        with st.spinner("Building workbook..."):
            xl=build_excel(branch,tills,treasury,journals,petty)
            fname=f"SUPERVISION_{branch.upper().replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.download_button("Download Excel Workbook",data=xl,file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success(f"Ready: {fname}")
